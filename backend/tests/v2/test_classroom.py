from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from core.class_policy import current_year_prefix
from db.models import ClassMembership, PlaySession, User, UserGrade
from tests.v2.test_teacher_grade_application import login_user, seed_student_user


def seed_teacher_user(db_session, *, login_id: str = "teacher01") -> User:
    user = seed_student_user(
        db_session,
        login_id=login_id,
        name="박선생",
        email=f"{login_id}@example.com",
    )
    user.grade = UserGrade.TEACHER
    db_session.commit()
    db_session.refresh(user)
    return user


def seed_completed_session(
    db_session,
    *,
    user_id: int,
    session_id: str,
    completed_at: datetime,
    history_score: int,
) -> PlaySession:
    session = PlaySession(
        id=session_id,
        user_id=user_id,
        scenario_id=None,
        ending_id=None,
        status="completed",
        choices_path=["A"],
        choices_history=[],
        history_score=history_score,
        final_stats={},
        character_name="세종대왕",
        scenario_title="기록된 세션",
        created_at=completed_at,
        completed_at=completed_at,
    )
    db_session.add(session)
    db_session.flush()
    return session


def load_excel_rows(content: bytes, sheet_name: str):
    workbook = load_workbook(BytesIO(content))
    sheet = workbook[sheet_name]
    return list(sheet.iter_rows(values_only=True))


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_create_class(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_01")
    token = login_user(client, "teacher_class_01")

    response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "2026년 4학년 1반", "entry_code_suffix": "4A1"},
    )

    assert response.status_code == 201
    data = response.json()
    assert data["name"] == "2026년 4학년 1반"
    assert data["entry_code"].endswith("4A1")
    assert data["entry_code"].startswith(current_year_prefix())
    assert data["is_active"] is True
    assert data["member_count"] == 0


@pytest.mark.usefixtures("jwt_env")
def test_student_cannot_create_class(client, db_session):
    seed_student_user(db_session, login_id="student_class_01")
    token = login_user(client, "student_class_01")

    response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "학생 클래스", "entry_code_suffix": "TEST"},
    )

    assert response.status_code == 403


@pytest.mark.usefixtures("jwt_env")
def test_teacher_lists_only_own_classes(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_02")
    token = login_user(client, "teacher_class_02")
    seed_teacher_user(db_session, login_id="teacher_class_03")
    other_token = login_user(client, "teacher_class_03")

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "내 클래스", "entry_code_suffix": "MINE"},
    )
    class_id = create_response.json()["id"]

    client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {other_token}"},
        json={"name": "다른 선생님 클래스", "entry_code_suffix": "OTHER"},
    )

    list_response = client.get(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert list_response.status_code == 200
    data = list_response.json()
    items = data["items"]
    assert len(items) == 1
    assert items[0]["id"] == class_id

    detail_response = client.get(
        f"/api/v2/classes/{class_id}",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert detail_response.status_code == 404


@pytest.mark.usefixtures("jwt_env")
def test_teacher_cannot_create_duplicate_entry_code(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_dup")
    token = login_user(client, "teacher_class_dup")

    first = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "1반", "entry_code_suffix": "DUP1"},
    )
    assert first.status_code == 201

    second = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "2반", "entry_code_suffix": "DUP1"},
    )
    assert second.status_code == 409


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_get_class_detail_with_members(client, db_session):
    from db.models import ClassMembership

    teacher = seed_teacher_user(db_session, login_id="teacher_class_detail")
    student = seed_student_user(db_session, login_id="student_in_class", name="김학생", nickname="학생1")
    token = login_user(client, "teacher_class_detail")

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "4학년 1반", "entry_code_suffix": "DETAIL"},
    )
    class_id = create_response.json()["id"]

    db_session.add(
        ClassMembership(
            class_id=class_id,
            user_id=student.id,
            is_active=True,
        )
    )
    db_session.commit()

    detail_response = client.get(
        f"/api/v2/classes/{class_id}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert detail_response.status_code == 200
    data = detail_response.json()
    assert data["name"] == "4학년 1반"
    assert data["member_count"] == 1
    assert len(data["members"]) == 1
    assert data["members"][0]["login_id"] == "student_in_class"
    assert data["members"][0]["name"] == "김학생"
    assert data["members"][0]["nickname"] == "학생1"
    assert data["members"][0]["is_active"] is True


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_get_class_activity_summary(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_summary")
    token = login_user(client, "teacher_class_summary")
    active_student_1 = seed_student_user(
        db_session,
        login_id="student_summary_1",
        name="학생1",
        nickname="별1",
    )
    active_student_2 = seed_student_user(
        db_session,
        login_id="student_summary_2",
        name="학생2",
        nickname="별2",
    )
    inactive_student = seed_student_user(
        db_session,
        login_id="student_summary_3",
        name="학생3",
        nickname="별3",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "활동 요약반", "entry_code_suffix": "SUM"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=active_student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=active_student_2.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=inactive_student.id, is_active=False),
        ]
    )
    db_session.flush()

    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="summary-session-1",
        completed_at=datetime(2026, 6, 10, 10, 0, 0),
        history_score=80,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="summary-session-2",
        completed_at=datetime(2026, 6, 20, 10, 0, 0),
        history_score=90,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_2.id,
        session_id="summary-session-3",
        completed_at=datetime(2026, 6, 25, 10, 0, 0),
        history_score=100,
    )
    seed_completed_session(
        db_session,
        user_id=inactive_student.id,
        session_id="summary-session-4",
        completed_at=datetime(2026, 6, 15, 10, 0, 0),
        history_score=50,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="summary-session-5",
        completed_at=datetime(2026, 5, 30, 10, 0, 0),
        history_score=70,
    )
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/activity-summary?date_from=2026-06-01&date_to=2026-06-30",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total_students"] == 2
    assert data["participating_students"] == 2
    assert data["completed_sessions"] == 3
    assert data["average_history_score"] == 90.0


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_list_class_student_sessions(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_student_sessions")
    token = login_user(client, "teacher_class_student_sessions")
    active_student_1 = seed_student_user(
        db_session,
        login_id="student_sessions_1",
        name="김학생",
        nickname="히어로",
    )
    active_student_2 = seed_student_user(
        db_session,
        login_id="student_sessions_2",
        name="이학생",
        nickname="용사",
    )
    inactive_student = seed_student_user(
        db_session,
        login_id="student_sessions_3",
        name="박학생",
        nickname="숨김",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "학생별 기록반", "entry_code_suffix": "SES"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=active_student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=active_student_2.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=inactive_student.id, is_active=False),
        ]
    )
    db_session.flush()

    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="student-session-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=85,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="student-session-2",
        completed_at=datetime(2026, 3, 8, 11, 20, 0),
        history_score=78,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_2.id,
        session_id="student-session-3",
        completed_at=datetime(2026, 3, 5, 9, 10, 0),
        history_score=92,
    )
    seed_completed_session(
        db_session,
        user_id=inactive_student.id,
        session_id="student-session-4",
        completed_at=datetime(2026, 3, 6, 9, 10, 0),
        history_score=50,
    )
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/student-sessions?page=1&page_size=10",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 2
    assert data["page"] == 1
    assert data["page_size"] == 10
    assert len(data["items"]) == 2

    student_by_login = {item["login_id"]: item for item in data["items"]}
    assert student_by_login["student_sessions_1"]["name"] == "김학생"
    assert student_by_login["student_sessions_1"]["completed_count"] == 2
    assert student_by_login["student_sessions_1"]["average_history_score"] == 81.5
    assert student_by_login["student_sessions_1"]["latest_completed_at"].startswith("2026-03-10")
    assert len(student_by_login["student_sessions_1"]["sessions"]) == 2
    assert student_by_login["student_sessions_1"]["sessions"][0]["history_score"] == 85

    assert student_by_login["student_sessions_2"]["completed_count"] == 1
    assert student_by_login["student_sessions_2"]["sessions"][0]["student_name"] == "이학생"


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_filter_class_student_sessions_by_name_and_date(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_student_sessions_filter")
    token = login_user(client, "teacher_class_student_sessions_filter")
    active_student_1 = seed_student_user(
        db_session,
        login_id="student_sessions_filter_1",
        name="김학생",
        nickname="히어로",
    )
    active_student_2 = seed_student_user(
        db_session,
        login_id="student_sessions_filter_2",
        name="이학생",
        nickname="용사",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "학생별 기록 필터반", "entry_code_suffix": "FLT"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=active_student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=active_student_2.id, is_active=True),
        ]
    )
    db_session.flush()

    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="student-session-filter-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=85,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="student-session-filter-2",
        completed_at=datetime(2026, 2, 10, 14, 30, 0),
        history_score=70,
    )
    seed_completed_session(
        db_session,
        user_id=active_student_2.id,
        session_id="student-session-filter-3",
        completed_at=datetime(2026, 3, 5, 9, 10, 0),
        history_score=92,
    )
    db_session.commit()

    name_response = client.get(
        f"/api/v2/classes/{class_id}/student-sessions?name=김학생",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert name_response.status_code == 200
    name_data = name_response.json()
    assert name_data["total"] == 1
    assert name_data["items"][0]["login_id"] == "student_sessions_filter_1"

    date_response = client.get(
        (
            f"/api/v2/classes/{class_id}/student-sessions"
            "?date_from=2026-03-01&date_to=2026-03-31"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert date_response.status_code == 200
    date_data = date_response.json()
    assert date_data["total"] == 2

    student_by_login = {item["login_id"]: item for item in date_data["items"]}
    assert student_by_login["student_sessions_filter_1"]["completed_count"] == 1
    assert len(student_by_login["student_sessions_filter_1"]["sessions"]) == 1
    assert student_by_login["student_sessions_filter_2"]["completed_count"] == 1
    assert student_by_login["student_sessions_filter_2"]["sessions"][0]["history_score"] == 92

    no_session_student = seed_student_user(
        db_session,
        login_id="student_sessions_filter_3",
        name="최학생",
        nickname="신입",
    )
    db_session.add(ClassMembership(class_id=class_id, user_id=no_session_student.id, is_active=True))
    db_session.commit()

    empty_session_response = client.get(
        (
            f"/api/v2/classes/{class_id}/student-sessions"
            "?date_from=2026-04-01&date_to=2026-04-30"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert empty_session_response.status_code == 200
    empty_session_data = empty_session_response.json()
    assert empty_session_data["total"] == 0
    assert empty_session_data["items"] == []


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_export_class_students(client, db_session):
    teacher = seed_teacher_user(db_session, login_id="teacher_export_students")
    token = login_user(client, teacher.login_id)
    student_1 = seed_student_user(
        db_session,
        login_id="student_export_1",
        name="김학생",
        nickname="히어로",
    )
    student_2 = seed_student_user(
        db_session,
        login_id="student_export_2",
        name="이학생",
        nickname="용사",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "학생 목록 export", "entry_code_suffix": "EXPSTU"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=student_2.id, is_active=True),
        ]
    )
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/students/export?nickname=히어로",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "class-" in response.headers["content-disposition"]

    rows = load_excel_rows(response.content, "학생 목록")
    assert rows[0] == ("아이디", "이름", "닉네임", "참여일")
    assert rows[1][0:3] == ("student_export_1", "김학생", "히어로")
    assert len(rows) == 2


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_export_class_student_sessions(client, db_session):
    teacher = seed_teacher_user(db_session, login_id="teacher_export_student_sessions")
    token = login_user(client, teacher.login_id)
    student = seed_student_user(
        db_session,
        login_id="student_export_sessions",
        name="김학생",
        nickname="히어로",
    )
    no_session_student = seed_student_user(
        db_session,
        login_id="student_export_sessions_2",
        name="이학생",
        nickname="용사",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "학생별 기록 export", "entry_code_suffix": "EXPSES"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=student.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=no_session_student.id, is_active=True),
        ]
    )
    db_session.flush()

    session = seed_completed_session(
        db_session,
        user_id=student.id,
        session_id="student-export-session-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=85,
    )
    session.character_name = "이황"
    session.scenario_title = "처음 만난 조선"
    db_session.commit()

    response = client.get(
        (
            f"/api/v2/classes/{class_id}/student-sessions/export"
            "?name=김학생&date_from=2026-03-01&date_to=2026-03-31"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    summary_rows = load_excel_rows(response.content, "학생별 요약")
    detail_rows = load_excel_rows(response.content, "세션 상세")

    assert summary_rows[0] == (
        "아이디",
        "이름",
        "닉네임",
        "참여일",
        "완료 건수",
        "평균 역사 점수",
        "최근 완료일",
    )
    assert summary_rows[1][0:3] == ("student_export_sessions", "김학생", "히어로")
    assert summary_rows[1][4] == 1
    assert detail_rows[0] == (
        "학생 아이디",
        "학생 이름",
        "학생 닉네임",
        "인물",
        "시나리오",
        "역사 점수",
        "완료일",
    )
    assert detail_rows[1][0:6] == (
        "student_export_sessions",
        "김학생",
        "히어로",
        "이황",
        "처음 만난 조선",
        85,
    )
    assert len(summary_rows) == 2
    assert len(detail_rows) == 2


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_list_class_play_sessions(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_records")
    token = login_user(client, "teacher_class_records")
    active_student_1 = seed_student_user(
        db_session,
        login_id="student_records_1",
        name="김학생",
        nickname="히어로",
    )
    active_student_2 = seed_student_user(
        db_session,
        login_id="student_records_2",
        name="이학생",
        nickname="용사",
    )
    inactive_student = seed_student_user(
        db_session,
        login_id="student_records_3",
        name="박학생",
        nickname="숨김",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "전체 기록반", "entry_code_suffix": "REC"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=active_student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=active_student_2.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=inactive_student.id, is_active=False),
        ]
    )
    db_session.flush()

    latest = seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="class-record-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=85,
    )
    latest.character_name = "이황"
    latest.scenario_title = "처음 만난 조선"

    older = seed_completed_session(
        db_session,
        user_id=active_student_2.id,
        session_id="class-record-2",
        completed_at=datetime(2026, 3, 8, 11, 20, 0),
        history_score=92,
    )
    older.character_name = "이이"
    older.scenario_title = "두 번째 조선"

    hidden = seed_completed_session(
        db_session,
        user_id=inactive_student.id,
        session_id="class-record-3",
        completed_at=datetime(2026, 3, 9, 9, 10, 0),
        history_score=50,
    )
    hidden.character_name = "세종대왕"
    hidden.scenario_title = "숨김 기록"
    db_session.commit()

    response = client.get(
        (
            f"/api/v2/classes/{class_id}/play-sessions"
            "?name=김학생&date_from=2026-03-01&date_to=2026-03-31&page=1&page_size=10"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["page"] == 1
    assert data["page_size"] == 10
    assert data["total"] == 1
    assert data["total_pages"] == 1
    assert len(data["items"]) == 1
    assert data["items"][0]["id"] == "class-record-1"
    assert data["items"][0]["student_login_id"] == "student_records_1"
    assert data["items"][0]["student_name"] == "김학생"
    assert data["items"][0]["student_nickname"] == "히어로"
    assert data["items"][0]["character_name"] == "이황"
    assert data["items"][0]["scenario_title"] == "처음 만난 조선"
    assert data["items"][0]["history_score"] == 85
    assert data["items"][0]["completed_at"].startswith("2026-03-10")

    all_response = client.get(
        f"/api/v2/classes/{class_id}/play-sessions?page=1&page_size=10",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert all_response.status_code == 200
    all_data = all_response.json()
    assert all_data["total"] == 2
    assert [item["id"] for item in all_data["items"]] == ["class-record-1", "class-record-2"]


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_export_class_play_sessions(client, db_session):
    teacher = seed_teacher_user(db_session, login_id="teacher_export_play_sessions")
    token = login_user(client, teacher.login_id)
    student = seed_student_user(
        db_session,
        login_id="student_export_records",
        name="김학생",
        nickname="히어로",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "전체 기록 export", "entry_code_suffix": "EXPREC"},
    )
    class_id = create_response.json()["id"]

    db_session.add(ClassMembership(class_id=class_id, user_id=student.id, is_active=True))
    db_session.flush()

    session = seed_completed_session(
        db_session,
        user_id=student.id,
        session_id="play-export-session-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=88,
    )
    session.character_name = "세종대왕"
    session.scenario_title = "조선의 시작"
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/play-sessions/export?nickname=히어로",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    rows = load_excel_rows(response.content, "전체 기록")
    assert rows[0] == (
        "학생 아이디",
        "학생 이름",
        "학생 닉네임",
        "인물",
        "시나리오",
        "역사 점수",
        "완료일",
    )
    assert rows[1][0:6] == (
        "student_export_records",
        "김학생",
        "히어로",
        "세종대왕",
        "조선의 시작",
        88,
    )


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_list_class_character_records(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_character_records")
    token = login_user(client, "teacher_class_character_records")
    active_student_1 = seed_student_user(
        db_session,
        login_id="student_character_records_1",
        name="김학생",
        email="student_character_records_1@example.com",
    )
    active_student_1.nickname = "히어로"
    active_student_2 = seed_student_user(
        db_session,
        login_id="student_character_records_2",
        name="이학생",
        email="student_character_records_2@example.com",
    )
    active_student_2.nickname = "용사"
    inactive_student = seed_student_user(
        db_session,
        login_id="student_character_records_3",
        name="박학생",
        email="student_character_records_3@example.com",
    )
    inactive_student.nickname = "숨김"
    db_session.commit()

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "인물 기록반", "entry_code_suffix": "CHR"},
    )
    class_id = create_response.json()["id"]

    db_session.add_all(
        [
            ClassMembership(class_id=class_id, user_id=active_student_1.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=active_student_2.id, is_active=True),
            ClassMembership(class_id=class_id, user_id=inactive_student.id, is_active=False),
        ]
    )
    db_session.flush()

    session_1 = seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="character-record-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=80,
    )
    session_1.character_name = "이황"
    session_1.scenario_title = "처음 만난 조선"

    session_2 = seed_completed_session(
        db_session,
        user_id=active_student_2.id,
        session_id="character-record-2",
        completed_at=datetime(2026, 3, 8, 11, 20, 0),
        history_score=90,
    )
    session_2.character_name = "이황"
    session_2.scenario_title = "처음 만난 조선"

    session_3 = seed_completed_session(
        db_session,
        user_id=active_student_1.id,
        session_id="character-record-3",
        completed_at=datetime(2026, 3, 9, 9, 10, 0),
        history_score=70,
    )
    session_3.character_name = "이이"
    session_3.scenario_title = "두 번째 조선"

    hidden = seed_completed_session(
        db_session,
        user_id=inactive_student.id,
        session_id="character-record-4",
        completed_at=datetime(2026, 3, 6, 9, 10, 0),
        history_score=50,
    )
    hidden.character_name = "세종대왕"
    hidden.scenario_title = "숨김 기록"
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/character-records?page=1&page_size=10",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["page"] == 1
    assert data["page_size"] == 10
    assert data["total"] == 2
    assert data["total_pages"] == 1
    assert len(data["items"]) == 2

    by_character = {item["character_name"]: item for item in data["items"]}
    assert by_character["이황"]["scenario_title"] == "처음 만난 조선"
    assert by_character["이황"]["average_history_score"] == 85.0
    assert by_character["이황"]["completed_student_count"] == 2
    assert by_character["이황"]["total_student_count"] == 2
    assert by_character["이이"]["average_history_score"] == 70.0
    assert by_character["이이"]["completed_student_count"] == 1

    filter_response = client.get(
        (
            f"/api/v2/classes/{class_id}/character-records"
            "?character_name=이황&date_from=2026-03-01&date_to=2026-03-31"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert filter_response.status_code == 200
    filter_data = filter_response.json()
    assert filter_data["total"] == 1
    assert filter_data["items"][0]["character_name"] == "이황"
    assert filter_data["items"][0]["completed_student_count"] == 2

    record_id = filter_data["items"][0]["id"]
    students_response = client.get(
        (
            f"/api/v2/classes/{class_id}/character-records/{record_id}/students"
            "?date_from=2026-03-01&date_to=2026-03-31"
        ),
        headers={"Authorization": f"Bearer {token}"},
    )
    assert students_response.status_code == 200
    students_data = students_response.json()
    assert len(students_data["items"]) == 2
    student_names = {item["name"] for item in students_data["items"]}
    assert student_names == {"김학생", "이학생"}


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_export_class_character_records(client, db_session):
    teacher = seed_teacher_user(db_session, login_id="teacher_export_character_records")
    token = login_user(client, teacher.login_id)
    student = seed_student_user(
        db_session,
        login_id="student_export_character",
        name="김학생",
        nickname="히어로",
    )

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "인물 기록 export", "entry_code_suffix": "EXPCHR"},
    )
    class_id = create_response.json()["id"]

    db_session.add(ClassMembership(class_id=class_id, user_id=student.id, is_active=True))
    db_session.flush()

    session = seed_completed_session(
        db_session,
        user_id=student.id,
        session_id="character-export-session-1",
        completed_at=datetime(2026, 3, 10, 14, 30, 0),
        history_score=91,
    )
    session.character_name = "이황"
    session.scenario_title = "처음 만난 조선"
    db_session.commit()

    response = client.get(
        f"/api/v2/classes/{class_id}/character-records/export?character_name=이황",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    record_rows = load_excel_rows(response.content, "인물 기록")
    student_rows = load_excel_rows(response.content, "완료 학생 상세")

    assert record_rows[0] == (
        "인물",
        "시나리오",
        "평균 역사 점수",
        "완료 학생 수",
        "전체 학생 수",
    )
    assert record_rows[1][0:4] == ("이황", "처음 만난 조선", 91, 1)
    assert student_rows[0] == (
        "인물",
        "시나리오",
        "학생 아이디",
        "학생 이름",
        "학생 닉네임",
        "역사 점수",
        "완료일",
    )
    assert student_rows[1][0:6] == (
        "이황",
        "처음 만난 조선",
        "student_export_character",
        "김학생",
        "히어로",
        91,
    )


@pytest.mark.usefixtures("jwt_env")
def test_teacher_can_update_and_delete_class(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_04")
    token = login_user(client, "teacher_class_04")

    create_response = client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "방과후 코딩반", "entry_code_suffix": "CODE"},
    )
    class_id = create_response.json()["id"]

    patch_response = client.patch(
        f"/api/v2/classes/{class_id}",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "2026 방과후 코딩반", "is_active": False},
    )
    assert patch_response.status_code == 200
    assert patch_response.json()["name"] == "2026 방과후 코딩반"
    assert patch_response.json()["is_active"] is False

    delete_response = client.delete(
        f"/api/v2/classes/{class_id}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert delete_response.status_code == 200

    list_response = client.get(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert list_response.json()["items"] == []


@pytest.mark.usefixtures("jwt_env")
def test_teacher_list_classes_pagination(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_page")
    token = login_user(client, "teacher_class_page")

    for index in range(3):
        client.post(
            "/api/v2/classes",
            headers={"Authorization": f"Bearer {token}"},
            json={"name": f"클래스 {index}", "entry_code_suffix": f"PG{index}"},
        )

    page1 = client.get(
        "/api/v2/classes?page=1&page_size=2",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert page1.status_code == 200
    data = page1.json()
    assert data["total"] == 3
    assert data["total_pages"] == 2
    assert len(data["items"]) == 2

    page2 = client.get(
        "/api/v2/classes?page=2&page_size=2",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert len(page2.json()["items"]) == 1


@pytest.mark.usefixtures("jwt_env")
def test_teacher_list_classes_filter_by_name(client, db_session):
    seed_teacher_user(db_session, login_id="teacher_class_search")
    token = login_user(client, "teacher_class_search")

    client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "4학년 1반", "entry_code_suffix": "SRCH1"},
    )
    client.post(
        "/api/v2/classes",
        headers={"Authorization": f"Bearer {token}"},
        json={"name": "방과후 코딩반", "entry_code_suffix": "SRCH2"},
    )

    response = client.get(
        "/api/v2/classes?name=코딩",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["total"] == 1
    assert data["items"][0]["name"] == "방과후 코딩반"
